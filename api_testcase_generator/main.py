# main.py
import os
import json
import argparse
from datetime import datetime
from typing import List, Tuple
from md_parser import parse_md_file
from testcase_generator import TestCaseGenerator
from models import TestCase, APIInfo


class TestCaseExporter:
    """测试用例导出器"""

    @staticmethod
    def to_markdown(test_cases: List[TestCase], api_info: APIInfo) -> str:
        """导出为Markdown格式"""
        lines = [
            f"# {api_info.name} 接口测试用例",
            f"\n> 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"> 接口名称: {api_info.action}",
            f"\n## 用例列表\n",
            "| 用例ID | 用例名称 | 所属模块 | 前置条件 | 步骤描述 | 预期结果 | 用例等级 |",
            "|--------|----------|----------|----------|----------|----------|----------|"
        ]

        for case in test_cases:
            precondition = case.precondition.replace('\n', '<br>')
            steps = case.steps.replace('\n', '<br>')
            expected = case.expected_result.replace('\n', '<br>')

            lines.append(
                f"| {case.case_id} | {case.case_name} | {case.module} | "
                f"{precondition} | {steps} | {expected} | {case.case_level.value} |"
            )

        return '\n'.join(lines)

    @staticmethod
    def to_json(test_cases: List[TestCase], api_info: APIInfo) -> str:
        """导出为JSON格式 (详细版, 输出到output目录)"""
        data = {
            "api_info": {
                "name": api_info.name,
                "action": api_info.action,
                "description": api_info.description
            },
            "generate_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "total_cases": len(test_cases),
            "test_cases": [case.to_dict() for case in test_cases]
        }
        return json.dumps(data, ensure_ascii=False, indent=2)

    @staticmethod
    def to_framework_json(test_cases: List[TestCase]) -> str:
        """导出为测试框架JSON格式 (输出到datas目录)"""
        data = {
            "test_cases": [case.to_framework_dict() for case in test_cases]
        }
        return json.dumps(data, ensure_ascii=False, indent=4)

    @staticmethod
    def to_excel(test_cases: List[TestCase], api_info: APIInfo, output_path: str) -> bool:
        """导出为Excel格式 (匹配模板11列)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            template_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                'templates', '测试用例模板.xlsx'
            )

            if os.path.exists(template_path):
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
                # 清除示例数据行 (保留表头)
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)
            else:
                # 模板不存在, 编程创建格式
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "模版"

                headers = [
                    "用例名称", "所属模块", "标签", "前置条件", "备注",
                    "步骤描述", "预期结果", "用例等级", "用例类型",
                    "是否支持自动化", "是否支持拨测"
                ]
                header_font = Font(name='宋体', size=14, bold=True)
                header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                ws.row_dimensions[1].height = 54
                ws.column_dimensions['B'].width = 30.83

            # 写入用例数据
            data_font = Font(name='等线', size=11)
            data_alignment = Alignment(vertical='center')

            for row_idx, case in enumerate(test_cases, 2):
                row_data = case.to_excel_row()
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment

            wb.save(output_path)
            return True
        except ImportError:
            print("  ⚠️ 请安装openpyxl: pip install openpyxl")
            return False
        except Exception as e:
            print(f"  ⚠️ Excel导出失败: {e}")
            return False

    @staticmethod
    def to_combined_excel(all_results: List[Tuple[APIInfo, List[TestCase]]], output_path: str) -> bool:
        """将所有接口的用例导出到一个Excel文件（多sheet）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            # 删除默认sheet
            wb.remove(wb.active)

            headers = [
                "用例名称", "所属模块", "标签", "前置条件", "备注",
                "步骤描述", "预期结果", "用例等级", "用例类型",
                "是否支持自动化", "是否支持拨测"
            ]
            header_font = Font(name='宋体', size=14, bold=True)
            header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            data_font = Font(name='等线', size=11)
            data_alignment = Alignment(vertical='center')

            def _write_sheet(ws, cases):
                # 写入表头
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                ws.row_dimensions[1].height = 54
                ws.column_dimensions['B'].width = 30.83
                # 写入数据
                for row_idx, case in enumerate(cases, 2):
                    row_data = case.to_excel_row()
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_alignment

            # 汇总sheet
            ws_summary = wb.create_sheet("汇总")
            summary_headers = ["接口名称", "Action", "用例总数", "P0", "P1", "P2", "P3", "P4"]
            for col_idx, header in enumerate(summary_headers, 1):
                cell = ws_summary.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            all_cases = []
            for idx, (api_info, test_cases) in enumerate(all_results):
                # 汇总行
                level_count = {}
                for case in test_cases:
                    level_prefix = case.case_level.value.split("-")[0]
                    level_count[level_prefix] = level_count.get(level_prefix, 0) + 1

                row = idx + 2
                ws_summary.cell(row=row, column=1, value=api_info.name)
                ws_summary.cell(row=row, column=2, value=api_info.action)
                ws_summary.cell(row=row, column=3, value=len(test_cases))
                ws_summary.cell(row=row, column=4, value=level_count.get("P0", 0))
                ws_summary.cell(row=row, column=5, value=level_count.get("P1", 0))
                ws_summary.cell(row=row, column=6, value=level_count.get("P2", 0))
                ws_summary.cell(row=row, column=7, value=level_count.get("P3", 0))
                ws_summary.cell(row=row, column=8, value=level_count.get("P4", 0))

                all_cases.extend(test_cases)

                # 每个接口单独sheet
                sheet_name = api_info.action[:31]
                ws_api = wb.create_sheet(sheet_name)
                _write_sheet(ws_api, test_cases)

            # 所有用例sheet
            ws_all = wb.create_sheet("所有用例")
            _write_sheet(ws_all, all_cases)

            wb.save(output_path)
            return True
        except ImportError:
            print("请安装openpyxl: pip install openpyxl")
            return False
        except Exception as e:
            print(f"Excel导出失败: {e}")
            return False


def get_md_files(docs_dir: str) -> List[str]:
    """获取目录下所有MD文件"""
    md_files = []

    if not os.path.exists(docs_dir):
        print(f"❌ 目录不存在: {docs_dir}")
        return md_files

    for filename in os.listdir(docs_dir):
        if filename.endswith('.md'):
            md_files.append(os.path.join(docs_dir, filename))

    # 按文件名排序
    md_files.sort()
    return md_files


def process_single_file(md_file_path: str, output_dir: str,
                        datas_dir: str = None, verbose: bool = True) -> Tuple[APIInfo, List[TestCase]]:
    """处理单个MD文件"""
    filename = os.path.basename(md_file_path)

    if verbose:
        print(f"\n{'='*60}")
        print(f"📄 处理文件: {filename}")
        print('='*60)

    # 1. 解析MD文件
    try:
        api_info = parse_md_file(md_file_path)
    except Exception as e:
        print(f"  ❌ 解析失败: {e}")
        return None, []

    if verbose:
        print(f"  接口名称: {api_info.name}")
        print(f"  Action: {api_info.action}")
        print(f"  请求参数: {len(api_info.request_params)} 个")

    # 2. 生成测试用例
    try:
        generator = TestCaseGenerator(api_info)
        test_cases = generator.generate()
    except Exception as e:
        print(f"  ❌ 生成用例失败: {e}")
        return api_info, []

    if verbose:
        print(f"  生成用例: {len(test_cases)} 条")

        # 统计等级分布
        level_count = {}
        for case in test_cases:
            level = case.case_level.value
            level_count[level] = level_count.get(level, 0) + 1
        level_str = ", ".join([f"{k}:{v}" for k, v in sorted(level_count.items())])
        print(f"  等级分布: {level_str}")

    # 3. 导出文件
    if output_dir:
        # Markdown → output/
        md_output = TestCaseExporter.to_markdown(test_cases, api_info)
        md_file = os.path.join(output_dir, f"{api_info.action}_testcases.md")
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write(md_output)

        # Excel → output/ (匹配模板11列)
        excel_file = os.path.join(output_dir, f"{api_info.action}_testcases.xlsx")
        TestCaseExporter.to_excel(test_cases, api_info, excel_file)

        if verbose:
            print(f"  ✅ 已导出到: {output_dir}")

    # 4. 框架JSON → datas/
    if datas_dir:
        if not os.path.exists(datas_dir):
            os.makedirs(datas_dir)
        json_output = TestCaseExporter.to_framework_json(test_cases)
        json_file = os.path.join(datas_dir, f"{api_info.action}.json")
        with open(json_file, 'w', encoding='utf-8') as f:
            f.write(json_output)
        if verbose:
            print(f"  ✅ 框架JSON已导出到: {json_file}")

    return api_info, test_cases


def process_all_files(docs_dir: str, output_dir: str, datas_dir: str = None):
    """批量处理所有MD文件"""
    print("\n" + "=" * 70)
    print("🚀 API接口测试用例批量生成工具")
    print("=" * 70)

    # 获取所有MD文件
    md_files = get_md_files(docs_dir)

    if not md_files:
        print(f"⚠️ 在 {docs_dir} 目录下未找到MD文件")
        return

    print(f"\n📂 文档目录: {docs_dir}")
    print(f"📁 输出目录: {output_dir}")
    if datas_dir:
        print(f"📁 数据目录: {datas_dir}")
    print(f"📝 发现 {len(md_files)} 个接口文档")

    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 处理每个文件
    all_results = []  # 存储所有结果
    success_count = 0
    fail_count = 0
    total_cases = 0

    for md_file in md_files:
        api_info, test_cases = process_single_file(md_file, output_dir, datas_dir)

        if api_info and test_cases:
            all_results.append((api_info, test_cases))
            success_count += 1
            total_cases += len(test_cases)
        else:
            fail_count += 1

    # 生成汇总Excel（所有接口在一个文件中）
    if all_results:
        combined_excel = os.path.join(output_dir, "_ALL_testcases.xlsx")
        if TestCaseExporter.to_combined_excel(all_results, combined_excel):
            print(f"\n📊 汇总Excel已生成: {combined_excel}")

    # 生成汇总报告
    print("\n" + "=" * 70)
    print("📈 生成完成 - 汇总报告")
    print("=" * 70)
    print(f"  处理文件数: {len(md_files)}")
    print(f"  成功: {success_count}")
    print(f"  失败: {fail_count}")
    print(f"  用例总数: {total_cases}")
    print(f"  输出目录: {output_dir}")

    # 列出生成的文件
    print(f"\n📁 生成的文件:")
    for api_info, test_cases in all_results:
        print(f"  - {api_info.action}_testcases.md/xlsx ({len(test_cases)}条)")
    print(f"  - _ALL_testcases.xlsx (汇总)")

    print("\n" + "=" * 70)
    print("✅ 全部完成!")
    print("=" * 70)


def main():
    """主函数 - 支持命令行参数"""
    parser = argparse.ArgumentParser(
        description='API接口测试用例生成工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 处理docs目录下的所有MD文件
  python main.py

  # 指定文档目录和输出目录
  python main.py -d ./api_docs -o ./test_output

  # 只处理单个文件
  python main.py -f ./docs/CreateTask.md
        """
    )

    parser.add_argument(
        '-d', '--docs-dir',
        default='./docs',
        help='接口文档目录 (默认: ./docs)'
    )

    parser.add_argument(
        '-o', '--output-dir',
        default='./output',
        help='输出目录 (默认: ./output)'
    )

    parser.add_argument(
        '-f', '--file',
        help='处理单个文件 (指定后忽略 -d 参数)'
    )

    parser.add_argument(
        '--datas-dir',
        default='../datas',
        help='框架JSON数据输出目录 (默认: ../datas)'
    )

    args = parser.parse_args()

    # 获取基础路径
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 解析datas_dir路径
    datas_dir = args.datas_dir if os.path.isabs(args.datas_dir) else os.path.join(base_dir, args.datas_dir)

    if args.file:
        # 处理单个文件
        file_path = args.file if os.path.isabs(args.file) else os.path.join(base_dir, args.file)
        output_dir = args.output_dir if os.path.isabs(args.output_dir) else os.path.join(base_dir, args.output_dir)

        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        if not os.path.exists(file_path):
            print(f"❌ 文件不存在: {file_path}")
            return

        process_single_file(file_path, output_dir, datas_dir)
        print(f"\n✅ 完成! 输出目录: {output_dir}")
        print(f"✅ 框架JSON目录: {datas_dir}")
    else:
        # 批量处理目录
        docs_dir = args.docs_dir if os.path.isabs(args.docs_dir) else os.path.join(base_dir, args.docs_dir)
        output_dir = args.output_dir if os.path.isabs(args.output_dir) else os.path.join(base_dir, args.output_dir)

        process_all_files(docs_dir, output_dir, datas_dir)


if __name__ == "__main__":
    main()
